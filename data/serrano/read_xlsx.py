# source: https://linuxhint.com/read-excel-file-python/
# Import openyxl module
import openpyxl
import numpy as np

# Define variable to load the wookbook
wookbook = openpyxl.load_workbook("Serrano_SIGG2021_PerceptualRatings.xlsx")

#print(wookbook.sheetnames)
# Define variable to read the active sheet:
#worksheet = wookbook.active
worksheet = wookbook["MedianRating_per_image"]

labels = []
# Iterate the loop to read the cell values
for i in range(0, worksheet.max_row):
    for col in worksheet.iter_cols(1,1):
        labels.append(col[i].value)
    if (i % 1000 == 0):
        print(f'[{i}/{worksheet.max_row}]')

print(labels)
np.save("samples_names", labels)
    